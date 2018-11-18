import openpyxl
import sys

def normalizeMatrix(sourceFileName, sourceSheetName, targetFileName, startCol, endCol, startRow, endRow):
	print("sourceFileName: %s"%sourceFileName)
	print("sourceSheetName: %s"%sourceSheetName)
	print("targetFileName: %s"%targetFileName)
	print("startCol: %s"%startCol)
	print("endCol: %s"%endCol)
	print("startRow: %s"%str(startRow))
	print("endRow: %s"%str(endRow))
	sourceWb = openpyxl.load_workbook(sourceFileName)
	sourceSheet = sourceWb.get_sheet_by_name(sourceSheetName)
	targetWb = openpyxl.Workbook()
	targetSheet = targetWb.get_sheet_by_name("Sheet")
	targetSheet.cell(row=1, column=1, value="A")
	targetSheet.cell(row=1, column=2, value="B")
	currentTargetRow = 2
	for row in range(int(startRow), int(endRow)):
		print("row: %s"%str(row))
		print("app %s"%str(sourceSheet[startCol+str(row)].value))
		listOfLocations  = []
		application = sourceSheet["A"+str(row)].value
		print("application: {}".format(application))
		for colNum in range(ord(startCol), ord(endCol)):
			col = chr(colNum)
			location = sourceSheet[col+"1"].value
			cell = sourceSheet[col+str(row)].value
			print("col+str(row): {}".format(col+str(row)))
			if(cell == "X"):
				listOfLocations.append(location)
		lengthOfList = len(listOfLocations)
		print("locations: {}".format(listOfLocations))
		print("length: {}".format(lengthOfList))
		for itemNum in range(0, lengthOfList):
			targetSheet.cell(row = currentTargetRow, column=1, value=application)
			targetSheet.cell(row = currentTargetRow, column=2, value=listOfLocations[itemNum])
			currentTargetRow += 1

	targetWb.save(targetFileName)
	print("Done")

if __name__ == "__main__":
	flattenSpreadsheet(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6], sys.argv[7])
